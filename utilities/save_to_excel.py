import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def save_data_to_excel(data, file_path, table_name):
    """
    Saves one or more profiles' data to Excel.
    - If the file doesn't exist, creates it with headers.
    - If it exists, appends new profiles only (avoids duplicates).

    Parameters:
        data (dict or list of dict): Single profile dict or list of profile dicts.
        file_path (str): Path to the Excel file.
    """
    # Normalize input to list of dicts
    try:
        if isinstance(data, dict):
            data = [data]
        elif not isinstance(data, list) or not all(isinstance(d, dict) for d in data):
            print("❌ Invalid data format. Must be dict or list of dicts.")
            return

        new_df = pd.DataFrame(data)

        # Determine data type based on columns
        if "content_id" in new_df.columns:
            unique_col = "content_id"
            data_type = "content-level"
        elif "Username" in new_df.columns and "content_id" not in new_df.columns:
            unique_col = "Username"
            data_type = "profile-level"
        else:
            print("❌ Cannot determine data type: missing 'Username' or 'Content ID'.")
            return

        # if not os.path.exists(file_path):
        #     # File doesn't exist: create new
        #     new_df.to_excel(file_path, index=False)
        #     print(f"✅ Created new file and saved {len(data)} profile(s).")
        # else:
        #     # File exists: read and append non-duplicate data
        #     existing_df = pd.read_excel(file_path)

        #     # Filter out profiles that already exist based on Username
        #     new_df = new_df[~new_df["Username"].isin(existing_df["Username"])]

        #     if new_df.empty:
        #         print("⚠️ All usernames already exist in file. Skipping save.")
        #         return

        #     updated_df = pd.concat([existing_df, new_df], ignore_index=True)
        #     updated_df.to_excel(file_path, index=False)
        #     print(f"✅ Appended {len(new_df)} new profile(s) to existing file.")

        if not os.path.exists(file_path):
            new_df.to_excel(file_path, index=False)
            print(f"✅ Created new file and saved {len(new_df)} {data_type} record(s).")
        else:
            existing_df = pd.read_excel(file_path)

            # Remove duplicates based on unique identifier
            new_df = new_df[~new_df[unique_col].isin(existing_df[unique_col])]

            if new_df.empty:
                print(
                    f"⚠️ All {data_type} records already exist in file. Skipping save."
                )
                return

            updated_df = pd.concat([existing_df, new_df], ignore_index=True)
            updated_df.to_excel(file_path, index=False)
            print(
                f"✅ Appended {len(new_df)} new {data_type} record(s) to existing file."
            )

    except Exception as e:
        print(f"❌ Error Saving the batch to excel: {e}")

    # Format Excel sheet as a styled table
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"

        # Remove any old table (if exists) to prevent duplication
        ws._tables.clear()

        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Set uniform column widths
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

        wb.save(file_path)
        print("📊 Excel sheet formatted as a table.")
    except Exception as e:
        print(f"⚠️ Failed to format Excel file as table: {e}")
